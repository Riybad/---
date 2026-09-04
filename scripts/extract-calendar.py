#!/usr/bin/env python3
"""يستخرج «الخطة الزمنية» من ملف تمكين إلى src/lib/calendar-data.ts

    pip install openpyxl && python3 scripts/extract-calendar.py

يقرأ ألوان الخلايا في شبكة الأشهر ويحوّلها إلى أحداث حسب دليل الألوان
أسفل الشبكة، ثم يضيف التاريخ الميلادي عبر تقويم أم القرى (Node).
لا يُشغَّل إلا إذا تغيّر ملف docs/الخطة-الزمنية.xlsx.
"""
import json
import subprocess
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "docs" / "الخطة-الزمنية.xlsx"
OUT = ROOT / "src" / "lib" / "calendar-data.ts"

# لون الخلفية في الملف -> مفتاح الحدث
COLORS = {
    "FFFFC000": "quran",       # الأيام القرآنية
    "FFFFFF00": "himma",       # أيام الهمة
    "FF4F6228": "intensive",   # دورة مكثفة
    "FF4F6128": "intensive",
    "FFD8E4BC": "session",     # دفعة 42 — وحدة التقسيم
    "FFD6E3BC": "session",
    "FFF2DBDB": "break",       # إجازة دراسية
    "FFF2DCDB": "break",
    "FFE4DFEC": "tamkeen",     # لقاء تمكين
    "FF31869B": "sleepover",   # مبيت
    "FF31859B": "sleepover",
    "FF980000": "forum",       # ملتقى
    "FF953734": "forum",
    "FF02FFFF": "recital",     # عرض المحفوظ
    "FF00FFFF": "recital",
    "FFB2A1C7": "summer",      # النادي الصيفي
    "FF76933B": "ramadan10",   # برنامج العشر الأواخر
    "FFD9D9D9": "weekend",
    "FFD8D8D8": "weekend",
    "FFFFFFFF": "none",
}

# صف بداية كل شهر في الورقة، واسمه ورقمه وسنته وفصله
MONTHS = [
    (18, "ربيع الأول", 1448, 3, "الفصل الأول"),
    (21, "ربيع الآخر", 1448, 4, "الفصل الأول"),
    (24, "جمادى الأولى", 1448, 5, "الفصل الأول"),
    (27, "جمادى الآخرة", 1448, 6, "الفصل الأول"),
    (30, "رجب", 1448, 7, "الفصل الأول"),
    (33, "شعبان", 1448, 8, "الفصل الثاني"),
    (36, "رمضان", 1448, 9, "الفصل الثاني"),
    (39, "شوال", 1448, 10, "الفصل الثاني"),
    (42, "ذو القعدة", 1448, 11, "الفصل الثاني"),
    (45, "ذو الحجة", 1448, 12, "الفصل الثاني"),
    (48, "محرم", 1449, 1, "الصيف"),
    (51, "صفر", 1449, 2, "الصيف"),
    (54, "ربيع الأول", 1449, 3, "الصيف"),
]

WEEKDAY_ROW = 16
FIRST_COL, LAST_COL = 9, 44  # I .. AR


def fill_of(cell) -> str:
    f = cell.fill
    if f is None or f.fill_type is None:
        return ""
    fg = f.fgColor
    if fg.type == "rgb" and fg.rgb != "00000000":
        return fg.rgb
    return ""


def event_of(ws, row: int, col: int) -> str:
    """الحدث يؤخذ من صف التاريخ أو الصفين تحته؛ العطلة الأسبوعية أضعف الأولويات."""
    weekend = ""
    for r in (row, row + 1, row + 2):
        key = COLORS.get(fill_of(ws.cell(r, col)), "")
        if key and key not in ("weekend", "none"):
            return key
        if key == "weekend":
            weekend = key
    return weekend or "none"


def main() -> int:
    if not SRC.exists():
        print(f"لم أجد الملف: {SRC}", file=sys.stderr)
        return 1

    ws = openpyxl.load_workbook(SRC).worksheets[0]
    # صف الترويسة يكرّر أيام الأسبوع عبر الشبكة — نأخذ الدورة الأولى فقط
    header = [ws.cell(WEEKDAY_ROW, c).value for c in range(FIRST_COL, LAST_COL + 1)]
    weekdays = header[:7]
    assert len(set(weekdays)) == 7, "ترويسة أيام الأسبوع غير متوقّعة"
    days = []

    for row, name, year, num, term in MONTHS:
        cols = [c for c in range(FIRST_COL, LAST_COL + 1) if ws.cell(row, c).value is not None]
        start, end = min(cols), max(cols)
        first = ws.cell(row, start).value
        # بعض خلايا الأرقام فيها أخطاء مطبعية في الملف، فنعتمد الترتيب لا القيمة
        base = first if isinstance(first, int) else 1
        for offset, col in enumerate(range(start, end + 1)):
            days.append({
                "hy": year,
                "hm": num,
                "hd": base + offset,
                "wd": weekdays.index(header[col - FIRST_COL]),
                "event": event_of(ws, row, col),
                "term": term,
            })

    # التاريخ الميلادي من تقويم أم القرى — متوفّر في Intl داخل Node
    node = subprocess.run(
        ["node", "-e", """
const days = JSON.parse(require("fs").readFileSync(0, "utf8"));
const fmt = new Intl.DateTimeFormat("en-u-ca-islamic-umalqura",
  { day: "numeric", month: "numeric", year: "numeric", timeZone: "UTC" });
const map = new Map();
for (let i = 0; i < 560; i++) {
  const dt = new Date(Date.UTC(2026, 7, 1) + i * 86400000);
  const p = Object.fromEntries(fmt.formatToParts(dt).map((x) => [x.type, x.value]));
  map.set(`${p.year}-${p.month}-${p.day}`, dt.toISOString().slice(0, 10));
}
for (const d of days) d.g = map.get(`${d.hy}-${d.hm}-${d.hd}`) ?? null;
process.stdout.write(JSON.stringify(days));
"""],
        input=json.dumps(days), capture_output=True, text=True, check=True,
    )
    days = json.loads(node.stdout)
    missing = [d for d in days if not d["g"]]
    if missing:
        print(f"تعذّر تحويل {len(missing)} يومًا إلى الميلادي", file=sys.stderr)
        return 1

    rows = ",\n".join(
        f'  [{d["hy"]},{d["hm"]},{d["hd"]},{d["wd"]},"{d["event"]}","{d["g"]}"]' for d in days
    )
    terms = json.dumps({f'{d["hy"]}-{d["hm"]}': d["term"] for d in days}, ensure_ascii=False, indent=2)
    keys = " | ".join(f'"{k}"' for k in dict.fromkeys(list(COLORS.values()) + ["none"]))

    OUT.write_text(f"""// مُولّد من ملف «الخطة الزمنية» — لا يُعدّل يدويًا (انظر scripts/extract-calendar.py)
export const WEEKDAYS = {json.dumps(weekdays, ensure_ascii=False)} as const;

export const HIJRI_MONTH_NAMES = ["محرم","صفر","ربيع الأول","ربيع الآخر","جمادى الأولى","جمادى الآخرة","رجب","شعبان","رمضان","شوال","ذو القعدة","ذو الحجة"] as const;

export type EventKey = {keys};

/** [سنة، شهر، يوم، رقم اليوم في الأسبوع، الحدث، التاريخ الميلادي] */
export const YEAR_DAYS: [number, number, number, number, EventKey, string][] = [
{rows},
];

export const TERM_OF_MONTH: Record<string, string> = {terms};
""", encoding="utf-8")
    print(f"كُتب {len(days)} يومًا إلى {OUT.relative_to(ROOT)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
